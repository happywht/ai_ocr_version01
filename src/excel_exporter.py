#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel导出工具
提供专业的Excel文件导出功能，支持多种格式和样式
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from typing import List, Dict, Any, Optional


class ExcelExporter:
    """Excel导出工具类"""

    def __init__(self):
        """初始化Excel导出工具"""
        # 定义样式
        self.setup_styles()

    def setup_styles(self):
        """设置Excel样式"""
        # 字体样式
        self.title_font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
        self.header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        self.content_font = Font(name='微软雅黑', size=11)
        self.number_font = Font(name='Arial', size=11)

        # 填充样式
        self.title_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
        self.header_fill = PatternFill(start_color='4A90E2', end_color='4A90E2', fill_type='solid')
        self.success_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
        self.warning_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
        self.error_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')

        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.thick_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        self.default_border = thin_border

        # 对齐样式
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        self.right_alignment = Alignment(horizontal='right', vertical='center')

    def export_single_invoice(self, file_path: str, invoice_data: Dict[str, Any],
                           format_type: str = "horizontal",
                           field_config: Optional[Dict[str, Any]] = None) -> bool:
        """
        导出单张发票数据

        Args:
            file_path: 导出文件路径
            invoice_data: 发票数据
            format_type: 导出格式 ("horizontal" 横向, "vertical" 纵向)
            field_config: 字段配置信息，用于动态生成表头

        Returns:
            导出是否成功
        """
        try:
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "发票识别结果"

            if format_type == "horizontal":
                self._create_horizontal_format(ws, invoice_data, field_config)
            else:
                self._create_vertical_format(ws, invoice_data, field_config)

            # 设置列宽
            self._auto_adjust_columns(ws)

            # 保存文件
            wb.save(file_path)
            return True

        except Exception as e:
            print(f"Excel导出失败: {str(e)}")
            return False

    def export_batch_invoices(self, file_path: str, invoices_data: List[Dict[str, Any]],
                              field_config: Optional[Dict[str, Any]] = None) -> bool:
        """
        批量导出发票数据

        Args:
            file_path: 导出文件路径
            invoices_data: 发票数据列表
            field_config: 字段配置信息，用于动态生成表头

        Returns:
            导出是否成功
        """
        try:
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "批量识别结果"

            # 获取动态字段列表
            dynamic_fields = self._get_dynamic_fields(invoices_data, field_config)

            # 创建表头 - 动态生成
            base_headers = ["序号", "图片路径", "处理时间", "解析方式", "AI置信度"]
            field_headers = [field_name for field_name in dynamic_fields]
            headers = base_headers + field_headers + ["识别状态"]

            # 设置表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.default_border
                cell.alignment = self.center_alignment

            # 填充数据
            for row, invoice in enumerate(invoices_data, 2):
                # 基础信息
                ws.cell(row=row, column=1, value=row-1)  # 序号
                ws.cell(row=row, column=2, value=invoice.get('图片路径', ''))
                ws.cell(row=row, column=3, value=invoice.get('处理时间', ''))
                ws.cell(row=row, column=4, value=invoice.get('解析方式', ''))
                ws.cell(row=row, column=5, value=invoice.get('AI置信度', ''))

                # 动态填充字段数据
                fields = invoice.get('提取字段', {})
                for col, field_name in enumerate(dynamic_fields, 6):
                    field_value = fields.get(field_name, '')
                    ws.cell(row=row, column=col, value=field_value)

                # 识别状态
                extracted_count = len([v for v in fields.values() if v])
                total_fields = len(fields)
                status = f"{extracted_count}/{total_fields}"
                status_col = len(headers)
                ws.cell(row=row, column=status_col, value=status)

                # 设置行样式
                for col in range(1, status_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = self.content_font
                    cell.border = self.default_border
                    cell.alignment = self.left_alignment

                    # 根据识别状态设置背景色
                    if col == status_col:  # 状态列
                        if extracted_count == total_fields:
                            cell.fill = self.success_fill
                        elif extracted_count >= total_fields * 0.7:
                            cell.fill = self.warning_fill
                        else:
                            cell.fill = self.error_fill
                    elif col >= 6 and col < status_col:  # 字段数据列
                        cell.font = self.number_font
                        cell.alignment = self.right_alignment

            # 动态设置列宽
            self._set_dynamic_column_widths(ws, len(base_headers), len(field_headers))

            # 保存文件
            wb.save(file_path)
            return True

        except Exception as e:
            print(f"批量Excel导出失败: {str(e)}")
            return False

    def _get_dynamic_fields(self, invoices_data: List[Dict[str, Any]],
                           field_config: Optional[Dict[str, Any]] = None) -> List[str]:
        """
        获取动态字段列表

        Args:
            invoices_data: 发票数据列表
            field_config: 字段配置信息

        Returns:
            排序后的字段名称列表
        """
        # 优先使用配置文件中的字段顺序
        if field_config and 'field_names' in field_config:
            return field_config['field_names']

        # 如果没有配置，从实际数据中提取字段
        all_fields = set()
        for invoice in invoices_data:
            fields = invoice.get('提取字段', {})
            all_fields.update(fields.keys())

        # 返回排序后的字段列表
        return sorted(list(all_fields))

    def _set_dynamic_column_widths(self, ws, base_headers_count: int, field_headers_count: int):
        """
        动态设置列宽

        Args:
            ws: 工作表对象
            base_headers_count: 基础表头数量
            field_headers_count: 字段表头数量
        """
        # 基础列宽
        base_widths = [8, 25, 20, 15, 12]  # 序号、图片路径、处理时间、解析方式、AI置信度

        # 字段列宽
        field_widths = [15] * field_headers_count  # 默认字段宽度

        # 状态列宽
        status_width = [12]

        # 组合所有列宽
        all_widths = base_widths + field_widths + status_width

        # 设置列宽
        for i, width in enumerate(all_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _create_horizontal_format(self, ws, invoice_data: Dict[str, Any],
                                field_config: Optional[Dict[str, Any]] = None):
        """创建横向格式 (序号/{字段_List})"""
        # 标题
        title_cell = ws.cell(row=1, column=1, value="识别结果")
        title_cell.font = self.title_font
        title_cell.fill = self.title_fill
        title_cell.border = self.thick_border
        title_cell.alignment = self.center_alignment
        ws.merge_cells('A1:D1')

        # 基本信息
        basic_info = [
            ("图片路径", invoice_data.get('图片路径', '')),
            ("处理时间", invoice_data.get('处理时间', '')),
            ("解析方式", invoice_data.get('解析方式', '')),
            ("AI置信度", invoice_data.get('AI置信度', ''))
        ]

        for i, (label, value) in enumerate(basic_info, 3):
            ws.cell(row=i, column=1, value=label).font = self.header_font
            ws.cell(row=i, column=1).fill = self.header_fill
            ws.cell(row=i, column=1).border = self.default_border
            ws.cell(row=i, column=1).alignment = self.center_alignment

            ws.cell(row=i, column=2, value=value).font = self.content_font
            ws.cell(row=i, column=2).border = self.default_border
            ws.cell(row=i, column=2).alignment = self.left_alignment
            ws.merge_cells(f'B{i}:D{i}')

        # 字段列表标题
        field_title_row = len(basic_info) + 4
        ws.cell(row=field_title_row, column=1, value="字段提取结果").font = self.title_font
        ws.cell(row=field_title_row, column=1).fill = self.title_fill
        ws.cell(row=field_title_row, column=1).border = self.thick_border
        ws.cell(row=field_title_row, column=1).alignment = self.center_alignment
        ws.merge_cells(f'A{field_title_row}:D{field_title_row}')

        # 字段列表表头
        headers = ["序号", "字段名称", "提取内容", "状态"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=field_title_row + 1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.default_border
            cell.alignment = self.center_alignment

        # 动态获取字段数据
        fields = invoice_data.get('提取字段', {})
        field_definitions = self._get_field_definitions(fields, field_config)

        for i, (field_name, description) in enumerate(field_definitions, field_title_row + 2):
            field_value = fields.get(field_name, '')
            status = "✅ 成功" if field_value else "❌ 未识别"

            # 序号
            ws.cell(row=i, column=1, value=i - field_title_row - 1).border = self.default_border
            ws.cell(row=i, column=1).alignment = self.center_alignment

            # 字段名称
            field_display_name = f"{field_name}\n({description})" if description else field_name
            ws.cell(row=i, column=2, value=field_display_name).border = self.default_border
            ws.cell(row=i, column=2).alignment = self.center_alignment

            # 提取内容
            content_cell = ws.cell(row=i, column=3, value=field_value)
            content_cell.border = self.default_border
            content_cell.alignment = self.left_alignment
            ws.merge_cells(f'C{i}:D{i}')

            # 状态列移除（因为已合并到内容列）
            if field_value:
                content_cell.fill = self.success_fill
            else:
                content_cell.fill = self.error_fill

    def _get_field_definitions(self, fields: Dict[str, Any],
                              field_config: Optional[Dict[str, Any]] = None) -> List[tuple]:
        """
        获取字段定义列表

        Args:
            fields: 字段数据
            field_config: 字段配置信息

        Returns:
            字段定义列表 [(field_name, description), ...]
        """
        field_definitions = []

        # 如果有配置文件，按配置顺序获取字段
        if field_config and 'fields' in field_config:
            config_fields = field_config['fields']
            for field_name in config_fields.keys():
                if field_name in fields:
                    field_info = config_fields[field_name]
                    description = field_info.get('description', '')
                    field_definitions.append((field_name, description))
        else:
            # 如果没有配置，按字段数据生成定义
            for field_name, field_value in fields.items():
                field_definitions.append((field_name, ''))

        return field_definitions

    def _create_vertical_format(self, ws, invoice_data: Dict[str, Any],
                              field_config: Optional[Dict[str, Any]] = None):
        """创建纵向格式"""
        # 标题
        title_cell = ws.cell(row=1, column=1, value="识别结果")
        title_cell.font = self.title_font
        title_cell.fill = self.title_fill
        title_cell.border = self.thick_border
        title_cell.alignment = self.center_alignment
        ws.merge_cells('A1:C1')

        # 基本信息
        row = 3
        basic_info = [
            ("图片路径", invoice_data.get('图片路径', '')),
            ("处理时间", invoice_data.get('处理时间', '')),
            ("解析方式", invoice_data.get('解析方式', '')),
            ("AI置信度", invoice_data.get('AI置信度', ''))
        ]

        for label, value in basic_info:
            ws.cell(row=row, column=1, value=label).font = self.header_font
            ws.cell(row=row, column=1).fill = self.header_fill
            ws.cell(row=row, column=1).border = self.default_border
            ws.cell(row=row, column=1).alignment = self.center_alignment

            ws.cell(row=row, column=2, value=value).font = self.content_font
            ws.cell(row=row, column=2).border = self.default_border
            ws.cell(row=row, column=2).alignment = self.left_alignment
            ws.merge_cells(f'B{row}:C{row}')
            row += 1

        # 字段详细信息
        row += 2
        ws.cell(row=row, column=1, value="字段详细信息").font = self.title_font
        ws.cell(row=row, column=1).fill = self.title_fill
        ws.cell(row=row, column=1).border = self.thick_border
        ws.cell(row=row, column=1).alignment = self.center_alignment
        ws.merge_cells(f'A{row}:C{row}')
        row += 1

        # 获取动态字段定义
        fields = invoice_data.get('提取字段', {})
        field_definitions = self._get_field_definitions(fields, field_config)

        for field_name, description in field_definitions:
            field_value = fields.get(field_name, '')

            # 字段名称
            field_display_name = f"{field_name}\n({description})" if description else field_name
            ws.cell(row=row, column=1, value=field_display_name).font = self.header_font
            ws.cell(row=row, column=1).fill = self.header_fill
            ws.cell(row=row, column=1).border = self.default_border
            ws.cell(row=row, column=1).alignment = self.center_alignment

            # 字段值
            value_cell = ws.cell(row=row, column=2, value=field_value or "未识别")
            value_cell.font = self.content_font
            value_cell.border = self.default_border
            value_cell.alignment = self.left_alignment
            ws.merge_cells(f'B{row}:C{row}')

            # 根据是否有值设置颜色
            if field_value:
                value_cell.fill = self.success_fill
            else:
                value_cell.fill = self.error_fill

            row += 1

    def _auto_adjust_columns(self, ws):
        """自动调整列宽"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_summary_sheet(self, wb, invoices_data: List[Dict[str, Any]]):
        """创建汇总表"""
        if not invoices_data:
            return

        ws = wb.create_sheet("数据汇总")

        # 统计信息
        total_invoices = len(invoices_data)
        successful_invoices = len([inv for inv in invoices_data
                                if len([v for v in inv.get('提取字段', {}).values() if v]) >= 4])

        # 创建汇总内容
        summary_data = [
            ("统计项目", "数值", "说明"),
            ("总发票数", total_invoices, "处理的发票总数"),
            ("成功识别", successful_invoices, "成功识别≥4个字段的发票"),
            ("识别成功率", f"{successful_invoices/total_invoices:.1%}" if total_invoices > 0 else "0%", "识别成功的比例"),
            ("导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "文件导出时间")
        ]

        for row, (item, value, desc) in enumerate(summary_data, 1):
            ws.cell(row=row, column=1, value=item).font = self.header_font
            ws.cell(row=row, column=1).fill = self.header_fill
            ws.cell(row=row, column=1).border = self.default_border

            ws.cell(row=row, column=2, value=value).font = self.content_font
            ws.cell(row=row, column=2).border = self.default_border

            ws.cell(row=row, column=3, value=desc).font = self.content_font
            ws.cell(row=row, column=3).border = self.default_border

        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30


def main():
    """测试Excel导出功能"""

    # 加载票据配置文件
    import json
    try:
        with open('peizhi001.json', 'r', encoding='utf-8') as f:
            peizhi_config = json.load(f)
    except FileNotFoundError:
        peizhi_config = None
        print("⚠️ 未找到peizhi001.json配置文件，使用默认配置")

    # 测试数据1: 票据识别
    test_invoice = {
        '图片路径': 'test_invoice.jpg',
        '处理时间': '2024-12-02 22:50:00',
        '解析方式': '🤖 AI智能解析',
        'AI置信度': 0.95,
        '提取字段': {
            '发票号码': '12345678',
            '开票日期': '2024-01-01',
            '销售方名称': '某某科技有限公司',
            '购买方名称': '某某贸易有限公司',
            '合计金额': '10600.00',
            '税额': '600.00',
            '合计': '10000.00'
        }
    }

    # 测试数据2: 图纸识别
    test_drawing = {
        '图片路径': 'test_drawing.png',
        '处理时间': '2024-12-16 10:30:00',
        '解析方式': '🤖 AI智能解析',
        'AI置信度': 0.88,
        '提取字段': {
            '项目名称': '某某科技园建设项目',
            '审定人': '张三',
            '审核人': '李四',
            '校核人': '王五',
            '设计人': '赵六',
            '绘图人': '孙七',
            '项目负责人': '周八',
            '专业负责人': '吴九',
            '项目编号': 'PROJ-2024-001',
            '图纸编号': 'DWG-2024-001-01',
            '设计阶段': '施工图',
            '专业': '建筑',
            '出图日期': '2024-12-01',
            '图纸比例': '1:100'
        }
    }

    exporter = ExcelExporter()

    # 测试单张发票导出 - 票据配置
    print("测试单张票据Excel导出（动态字段）...")
    if exporter.export_single_invoice("test_single_peizhi_invoice.xlsx", test_invoice, "horizontal", peizhi_config):
        print("✅ 单张票据导出成功")
    else:
        print("❌ 单张票据导出失败")

    # 测试单张图纸导出
    print("\n测试单张图纸Excel导出（动态字段）...")
    if exporter.export_single_invoice("test_single_drawing.xlsx", test_drawing, "horizontal"):
        print("✅ 单张图纸导出成功")
    else:
        print("❌ 单张图纸导出失败")

    # 测试批量导出 - 混合数据
    print("\n测试批量Excel导出（动态字段）...")
    test_invoices = [test_invoice, test_drawing, test_invoice]  # 混合票据和图纸数据
    if exporter.export_batch_invoices("test_dynamic_batch.xlsx", test_invoices):
        print("✅ 批量动态导出成功")
    else:
        print("❌ 批量动态导出失败")

    print("\n🎉 Excel动态字段导出功能测试完成！")
    print("📁 生成的文件:")
    print("   - test_single_peizhi_invoice.xlsx (票据格式)")
    print("   - test_single_drawing.xlsx (图纸格式)")
    print("   - test_dynamic_batch.xlsx (批量动态字段)")


if __name__ == "__main__":
    main()